#!/usr/bin/env python3
"""
🔍 미매칭 상품 스캐너 + 마스터 시트 자동 추가

이 스크립트는:
1. 네이버 API에서 상품 매출 데이터 가져옴 (지정 기간)
2. 마스터 시트와 매칭 시도
3. 미매칭 상품을 마스터 시트에 자동 추가:
   - 마스터상품명: API에서 받은 raw 상품명
   - 옵션코드1: API의 productId
   - Brand: 브랜드명
   - Type/Line: "?" (수동으로 채울 부분)
   - 원가: 비워둠 (수동으로 채울 부분)

[사용법]
  python scan_unmapped.py pacsafe                 # 팩세이프, 최근 30일
  python scan_unmapped.py president               # 프레지던트, 최근 30일
  python scan_unmapped.py pacsafe --days 60       # 최근 60일
  python scan_unmapped.py president --days 90     # 최근 90일

실행 후 마스터 엑셀이 백업되고 새 행이 추가됨.
사용자는 엑셀 열어서 원가/Type/Line만 채우면 끝.
"""
import os
import sys
import shutil
from datetime import datetime, timedelta
from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# report_core에서 필요한 함수만 import
from report_core import (
    get_token_and_channel,
    collect_data,
    load_product_catalog,
    match_product,
    KST,
)


BRAND_CONFIG = {
    "pacsafe": {
        "prefix_kr": "팩세이프",
        "prefix_en": "PACSAFE",
        "id_env": "NAVER_CLIENT_ID_PACSAFE",
        "secret_env": "NAVER_CLIENT_SECRET_PACSAFE",
        "catalog_file": "네이버_팩세이프_마스터_상품명.xlsx",
    },
    "president": {
        "prefix_kr": "프레지던트",
        "prefix_en": "PRSIDENT",  # 시트의 Brand 값 그대로
        "id_env": "NAVER_CLIENT_ID_PRESIDENT",
        "secret_env": "NAVER_CLIENT_SECRET_PRESIDENT",
        "catalog_file": "네이버_프레지던트_마스터_상품명.xlsx",
    },
}


def collect_unmapped(brand_key, days):
    """N일치 데이터를 일별로 수집해서 미매칭 상품 dict 반환"""
    cfg = BRAND_CONFIG[brand_key]
    naver_id = os.environ.get(cfg["id_env"])
    naver_secret = os.environ.get(cfg["secret_env"])

    if not naver_id or not naver_secret:
        print(f"❌ 환경변수 누락: {cfg['id_env']} / {cfg['secret_env']}")
        sys.exit(1)

    # 카탈로그 로드
    script_dir = os.path.dirname(os.path.abspath(__file__))
    catalog_path = os.path.join(script_dir, cfg["catalog_file"])
    if not os.path.exists(catalog_path):
        print(f"❌ 마스터 시트 없음: {catalog_path}")
        sys.exit(1)

    by_code, by_master, all_items = load_product_catalog(catalog_path)
    print(f"📚 카탈로그: {len(all_items)}개 상품 / {len(by_code)}개 옵션코드\n")

    # 토큰 발급
    token, cno = get_token_and_channel(naver_id, naver_secret)
    if not token or not cno:
        print("❌ 토큰 발급 실패")
        sys.exit(1)
    print(f"🔑 채널: {cno}\n")

    # N일치 데이터 일별 수집
    today = datetime.now(KST).date()
    unmapped = {}  # productId → {raw_name, total_purchases, total_revenue}

    for offset in range(1, days + 1):
        d = today - timedelta(days=offset)
        ds = d.strftime("%Y-%m-%d")
        try:
            data = collect_data(token, cno, ds, ds)
            sales = data.get("sales_p", {}).get("productUnitReport", []) or []

            day_unmapped = 0
            for p in sales:
                raw_name = p.get("productName", "")
                raw_id = (p.get("channelProductNo") or
                          p.get("originProductNo") or
                          p.get("productNo") or
                          p.get("productId") or
                          p.get("productOptionId") or
                          p.get("productOptionNo"))
                match = match_product(raw_name, raw_id, by_code, by_master, all_items)
                if not match:
                    pid = str(raw_id) if raw_id else f"NO_ID_{raw_name[:30]}"
                    if pid not in unmapped:
                        unmapped[pid] = {
                            "productId": str(raw_id) if raw_id else "",
                            "productName": raw_name or "",
                            "total_purchases": 0,
                            "total_revenue": 0,
                        }
                    unmapped[pid]["total_purchases"] += int(p.get("numPurchases", 0) or 0)
                    unmapped[pid]["total_revenue"] += int(p.get("payAmount", 0) or 0)
                    day_unmapped += 1

            print(f"  {ds}: 전체 {len(sales)}개 / 미매칭 {day_unmapped}개")
        except Exception as e:
            print(f"  {ds}: 에러 - {e}")

    return unmapped, catalog_path, cfg


def add_to_master(unmapped, catalog_path, cfg):
    """미매칭 상품을 마스터 엑셀에 자동 추가"""
    if not unmapped:
        print("\n✅ 미매칭 상품 없음. 모두 매핑되어 있어요!")
        return

    # 백업
    backup_path = catalog_path.replace(".xlsx", f"_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    shutil.copy2(catalog_path, backup_path)
    print(f"\n📦 백업: {os.path.basename(backup_path)}")

    # 엑셀 열기
    wb = load_workbook(catalog_path)
    ws = wb["마스터상품명"]

    # 마지막 데이터 행 찾기 (셀에 값 있는 마지막 행)
    last_data_row = 0
    for i in range(1, ws.max_row + 1):
        if ws.cell(row=i, column=1).value:
            last_data_row = i

    # 매출 큰 순 정렬
    sorted_items = sorted(unmapped.values(),
                          key=lambda x: x["total_revenue"], reverse=True)

    # 추가
    added = 0
    for item in sorted_items:
        last_data_row += 1
        # 컬럼 (1-indexed):
        # 1: 마스터상품명, 2: 원가, 3-12: 옵션코드1~10, 13: SKU
        # 14: Brand, 15: Type, 16: Line
        ws.cell(row=last_data_row, column=1, value=item["productName"][:80])
        # 원가는 비워둠 (수동 입력)
        if item["productId"]:
            try:
                ws.cell(row=last_data_row, column=3, value=int(item["productId"]))
            except (ValueError, TypeError):
                ws.cell(row=last_data_row, column=3, value=item["productId"])
        # SKU 비움
        ws.cell(row=last_data_row, column=14, value=cfg["prefix_en"])
        ws.cell(row=last_data_row, column=15, value="?")  # Type
        ws.cell(row=last_data_row, column=16, value="?")  # Line
        added += 1

    wb.save(catalog_path)
    print(f"\n✅ {added}개 행 마스터 시트에 추가됨")
    print(f"   파일: {os.path.basename(catalog_path)}")


def print_summary(unmapped):
    """미매칭 요약 출력"""
    if not unmapped:
        return
    print(f"\n{'='*60}")
    print(f"📋 미매칭 상품 요약 (총 {len(unmapped)}개, 매출 큰 순 TOP 20)")
    print(f"{'='*60}")
    print(f"{'productId':<14} {'판매':>4} {'매출':>12}  상품명")
    print("-" * 60)
    sorted_items = sorted(unmapped.values(),
                          key=lambda x: x["total_revenue"], reverse=True)
    for item in sorted_items[:20]:
        pid = item["productId"][:13]
        cnt = item["total_purchases"]
        rev = item["total_revenue"]
        name = item["productName"][:50]
        print(f"{pid:<14} {cnt:>4} {rev:>12,}  {name}")
    if len(unmapped) > 20:
        print(f"... 외 {len(unmapped) - 20}개")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    brand_key = sys.argv[1].lower()
    if brand_key not in BRAND_CONFIG:
        print(f"❌ 브랜드는 'pacsafe' 또는 'president' 만 가능. 입력: {brand_key}")
        sys.exit(1)

    # --days 파싱
    days = 30
    for i, arg in enumerate(sys.argv):
        if arg == "--days" and i + 1 < len(sys.argv):
            try:
                days = int(sys.argv[i + 1])
            except ValueError:
                pass

    cfg = BRAND_CONFIG[brand_key]
    print(f"🔍 {cfg['prefix_kr']} 미매칭 상품 스캔")
    print(f"   기간: 최근 {days}일\n")

    unmapped, catalog_path, cfg = collect_unmapped(brand_key, days)

    print_summary(unmapped)

    if unmapped:
        ans = input(f"\n👉 위 {len(unmapped)}개를 마스터 시트에 추가할까요? [y/N]: ")
        if ans.strip().lower() == "y":
            add_to_master(unmapped, catalog_path, cfg)
            print(f"\n다음 단계: 엑셀 열어서 추가된 행의 Type/Line/원가 수동 입력")
        else:
            print("취소됨.")


if __name__ == "__main__":
    main()
