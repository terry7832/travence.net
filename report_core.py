#!/usr/bin/env python3
"""
📊 BIZRAW v2.0 - 공통 리포트 엔진
- 모든 브랜드(팩세이프/프레지던트 등)가 공유하는 핵심 로직
- 데이터 수집 / 분석 / AI / HTML 생성 / 메일 발송
"""
import os
import sys
import time
import json
import base64
import logging
import smtplib
from datetime import datetime, timedelta, timezone
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import bcrypt
import requests
import anthropic

# =========================================================
# 공통 설정
# =========================================================
KST = timezone(timedelta(hours=9))
BASE_URL = "https://api.commerce.naver.com/external"
CLAUDE_MODEL = "claude-opus-4-7"  # Opus 4.7 (최신 플래그십)

# 네이버 커머스 API 엔드포인트
PATHS = {
    # 기본
    "sales_p":         "/v1/bizdata-stats/channels/{cno}/sales/product/detail",
    "mkt_all":         "/v1/bizdata-stats/channels/{cno}/marketing/all/detail",
    "keyword":         "/v1/bizdata-stats/channels/{cno}/marketing/search/keyword",
    "delivery":        "/v1/bizdata-stats/channels/{cno}/sales/delivery/detail",
    "hourly_s":        "/v1/bizdata-stats/channels/{cno}/sales/hourly/detail",
    # v2.0 신규 추가
    "category_sales":  "/v1/bizdata-stats/channels/{cno}/sales/product-marketing/category",
    "product_keyword": "/v1/bizdata-stats/channels/{cno}/sales/product-search/keyword-by-product",
    "customer_status": "/v1/customer-data/customer-status/channels/{cno}/statistics",
    "repurchase":      "/v1/customer-data/repurchase/account/statistics",
}

# =========================================================
# 팩세이프 정리된 상품 카탈로그 (53개) - AI 매핑 기준
# Raw 상품명을 이 리스트로 매핑
# =========================================================

# =========================================================
# 마스터 상품 카탈로그 (엑셀 파일에서 동적 로드)
# =========================================================
def load_product_catalog(excel_path):
    """마스터 엑셀 → 매핑 dict 생성"""
    by_code = {}     # 옵션코드 → 카탈로그
    by_master = {}   # 마스터상품명 → 카탈로그
    all_items = []   # 키워드 매칭용

    if not excel_path or not os.path.exists(excel_path):
        logger.warning(f"⚠️  카탈로그 파일 없음: {excel_path}")
        logger.warning(f"   상품명 매핑 비활성화 — raw 상품명 그대로 표시됩니다.")
        return by_code, by_master, all_items

    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, data_only=True)
        ws = wb['마스터상품명']

        for row in ws.iter_rows(min_row=3, values_only=True):
            name = row[0]
            if not name or '▣' in str(name):
                continue
            item = {
                'master_name': str(name).strip(),
                'cost': float(row[1] or 0),  # 원가
                'type': str(row[14] or '').strip() if row[14] else '',
                'line': str(row[15] or '').strip() if row[15] else '',
                'sku': str(row[12] or '').strip() if row[12] else '',
            }
            # 옵션코드 1~10 → 카탈로그 매핑
            for c in row[2:12]:
                if c is None:
                    continue
                code_str = str(int(c)) if isinstance(c, float) else str(c).strip()
                if code_str and code_str != 'None':
                    by_code[code_str] = item
            by_master[item['master_name']] = item
            all_items.append(item)

        logger.info(f"📚 카탈로그 로드: {len(all_items)}개 상품 / {len(by_code)}개 옵션코드")
    except ImportError:
        logger.error("❌ openpyxl 미설치. pip install openpyxl 필요")
    except Exception as e:
        logger.error(f"❌ 카탈로그 로드 실패: {e}")

    return by_code, by_master, all_items


def match_product(raw_name, raw_id, by_code, by_master, all_items):
    """raw 상품 → 마스터 매칭 (3단계: 옵션코드 → 상품명 포함 → 키워드)"""
    if not all_items:
        return None

    # 1. 옵션코드 직접 매칭 (가장 정확)
    if raw_id:
        rid = str(raw_id).strip()
        if rid.endswith('.0'):
            rid = rid[:-2]
        if rid in by_code:
            return by_code[rid]

    # 2. 마스터상품명이 raw에 포함
    if raw_name:
        raw_l = raw_name.lower()
        # 긴 마스터부터 우선 매칭 (특이성 높은 게 먼저)
        for master in sorted(by_master.keys(), key=len, reverse=True):
            if master.lower() in raw_l:
                return by_master[master]

    # 3. 키워드 매칭 (마스터의 모든 단어 70% 이상 raw에 포함)
    if raw_name:
        raw_l = raw_name.lower()
        best = None
        best_score = 0.7
        for item in all_items:
            keywords = item['master_name'].split()
            if not keywords:
                continue
            matched = sum(1 for kw in keywords if kw.lower() in raw_l)
            score = matched / len(keywords)
            if score >= best_score:
                best_score = score
                best = item
        return best

    return None


logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)


# =========================================================
# 유틸
# =========================================================
def _show_margin(stats):
    """마진 표시 여부: 브랜드 설정(show_margin) + 원가 신뢰성(cost_reliable) 둘 다 만족해야 표시"""
    return stats.get("show_margin", True) and stats.get("cost_reliable", True)


def format_curr(val):
    if not val:
        return "0"
    if val >= 100000000:
        return f"{val/100000000:.1f}억"
    elif val >= 10000:
        return f"{val/10000:,.0f}만"
    return f"{val:,.0f}"


def pct_change(curr, prev):
    """변화율 계산 (%, prev=0이면 None 반환)"""
    if not prev or prev == 0:
        return None
    return (curr - prev) / prev * 100


def get_delta_chip(curr, prev, is_percent=False, inverse=False):
    """변화량을 칩 HTML로 반환"""
    if not prev or prev == 0:
        return "<span style='color:#94a3b8; font-size:11px;'>-</span>"
    diff = curr - prev
    if is_percent:
        val = diff
        txt = f"{val:+.1f}%p"
        is_up = val > 0
    else:
        val = (diff / prev) * 100
        txt = f"{val:+.1f}%"
        is_up = val > 0

    # inverse=True: 증가가 나쁜 지표 (광고비, 반품 등)
    if inverse:
        bg, text = ("#fef2f2", "#ef4444") if is_up else ("#f0fdf4", "#10b981")
    else:
        bg, text = ("#dcfce7", "#15803d") if is_up else ("#fee2e2", "#b91c1c")
    icon = "▲" if is_up else "▼"
    return (f"<span style='display:inline-flex; align-items:center; gap:4px; "
            f"background:{bg}; color:{text}; padding:3px 8px; border-radius:12px; "
            f"font-size:11px; font-weight:700;'>{icon} {txt}</span>")


# =========================================================
# 1. 데이터 수집
# =========================================================
def get_token_and_channel(client_id, client_secret):
    """네이버 커머스 OAuth 토큰 + 채널 번호 조회"""
    ts = str(int((time.time() - 3) * 1000))
    pwd = f"{client_id}_{ts}"
    hashed = bcrypt.hashpw(pwd.encode("utf-8"), client_secret.encode("utf-8"))
    sign = base64.b64encode(hashed).decode("utf-8")

    try:
        res = requests.post(f"{BASE_URL}/v1/oauth2/token", data={
            "client_id": client_id, "timestamp": ts,
            "grant_type": "client_credentials", "client_secret_sign": sign, "type": "SELF"
        }, timeout=10)
        token = res.json().get("access_token")

        res = requests.get(f"{BASE_URL}/v1/seller/channels",
                           headers={"Authorization": f"Bearer {token}"}, timeout=10)
        raw = res.json()
        channels = raw if isinstance(raw, list) else raw.get("data", raw.get("contents", []))

        for ch in channels:
            if ch.get("channelType") == "STOREFARM" and ch.get("channelNo"):
                return token, ch["channelNo"]
        return token, channels[0].get("channelNo") if channels else None
    except Exception as e:
        logger.error(f"토큰/채널 조회 실패: {e}")
        return None, None


def collect_data(token, cno, start_date, end_date):
    """모든 API 데이터 수집 (실패한 API는 빈 dict 반환)"""
    data = {}
    headers = {"Authorization": f"Bearer {token}"}
    params_default = {"startDate": start_date, "endDate": end_date}

    for key, path in PATHS.items():
        try:
            url = f"{BASE_URL}{path.format(cno=cno)}" if "{cno}" in path else f"{BASE_URL}{path}"
            res = requests.get(url, headers=headers, params=params_default, timeout=15)
            if res.status_code == 200:
                data[key] = res.json()
                logger.info(f"  ✓ {key}")
            else:
                logger.warning(f"  ✗ {key} (HTTP {res.status_code})")
                data[key] = {}
        except Exception as e:
            logger.warning(f"  ✗ {key} 실패: {e}")
            data[key] = {}
    return data


# =========================================================
# 2. 데이터 분석
# =========================================================
def _extract_rows(raw, *possible_keys):
    """API 응답에서 row list를 안전하게 추출 (list/dict/이상한 응답 모두 대응)"""
    if isinstance(raw, list):
        return raw
    if isinstance(raw, dict):
        for key in possible_keys:
            val = raw.get(key)
            if val and isinstance(val, list):
                return val
    return []


def _guess_category_from_raw(raw_name):
    """raw 상품명에서 키워드로 카테고리 추측 (8위 밖 상품용)"""
    if not raw_name:
        return "Misc"
    name_lower = raw_name.lower()
    # 키워드 매칭 (한글/영어)
    if any(kw in name_lower for kw in ["wallet", "지갑", "rfidsleeve", "rfidsafe", "card sleeve", "카드"]):
        return "Wallet"
    if any(kw in name_lower for kw in ["sling", "슬링", "hip pack", "힙팩", "힙색", "waist pack"]):
        return "Sling"
    if any(kw in name_lower for kw in ["backpack", "백팩", "carry-on", "캐리온", "travel pack"]):
        return "Backpack"
    if any(kw in name_lower for kw in ["crossbody", "크로스바디", "크로스백"]):
        return "Crossbody"
    if any(kw in name_lower for kw in ["tote", "토트"]):
        return "Tote"
    return "Misc"


def analyze_data(data, catalog=None):
    """수집한 데이터를 stats 딕셔너리로 정리 + 마스터 카탈로그 매핑"""
    by_code, by_master, all_items = catalog or ({}, {}, [])

    stats = {
        "total_revenue": 0, "total_purchases": 0, "total_refund": 0, "total_cancel": 0, "total_cost": 0,
        "products": [], "channels": [], "keywords": [], "hourly_sales": [],
        "delivery": {}, "devices": {},
        # v2.0 신규
        "categories": [],          # [{name, revenue, purchases, share}]
        "lines": [],               # [{name, revenue, purchases, share}] - 신규 추가
        "product_keywords": [],    # [{product, keywords:[...]}]
        "customer_new": 0,         # 신규 구매자 수
        "customer_returning": 0,   # 기존 구매자 수
        "repurchase_rate": 0,      # 재구매율 (%)
        "total_cogs": 0,           # 매출원가 - 신규 추가
        "total_gp": 0,             # 매출총이익 - 신규 추가
        "margin_rate": 0,          # 마진율 (%) - 신규 추가
    }

    # --- 상품별 매출 + 마스터 매핑 + 원가/이익 계산
    sales_list = data.get('sales_p', {}).get('productUnitReport', []) or []

    # 🔍 디버그: 첫 상품의 모든 필드 출력 (ID 필드 확인용)
    if sales_list and all_items:
        first = sales_list[0]
        logger.info(f"  🔍 [디버그] 첫 상품 raw 필드: {list(first.keys())}")
        # ID 후보 필드 값 출력
        id_fields = ['channelProductNo', 'originProductNo', 'productNo', 'productId',
                     'productOptionId', 'productOptionNo', 'optionCode', 'productCode']
        for f in id_fields:
            if f in first:
                logger.info(f"     {f} = {first[f]}")
        logger.info(f"     productName = {first.get('productName', '')[:80]}")

    matched_count = 0
    failed_samples = []
    failed_all = []  # 매핑 실패 전체 리스트 (CSV용)
    total_gp = 0  # 전체 매출총이익
    total_cogs = 0  # 전체 원가 (매출원가)
    for p in sales_list:
        raw_name = p.get('productName', '')
        raw_id = (p.get('channelProductNo') or
                  p.get('originProductNo') or
                  p.get('productNo') or
                  p.get('productId') or
                  p.get('productOptionId') or
                  p.get('productOptionNo'))
        match = match_product(raw_name, raw_id, by_code, by_master, all_items)
        cnt = int(p.get('numPurchases', 0) or 0)
        amt = p.get('payAmount', 0) or 0
        if match:
            p['_master_name'] = match['master_name']
            p['_category'] = match['type'] or '미분류'
            p['_line'] = match['line'] or '?'
            p['_unit_cost'] = match['cost']  # 단가 원가
            p['_cogs'] = match['cost'] * cnt  # 매출원가
            p['_gp'] = amt - p['_cogs']  # 매출총이익
            p['_margin_rate'] = (p['_gp'] / amt * 100) if amt > 0 else 0
            total_cogs += p['_cogs']
            total_gp += p['_gp']
            matched_count += 1
        else:
            p['_master_name'] = raw_name[:60] if raw_name else '?'
            p['_category'] = '미분류'
            p['_line'] = '?'
            p['_unit_cost'] = 0
            p['_cogs'] = 0
            p['_gp'] = 0
            p['_margin_rate'] = 0
            if len(failed_samples) < 5:
                failed_samples.append(f"  ✗ '{raw_name[:60]}' (id={raw_id})")
            # CSV용 전체 수집
            failed_all.append({
                'productId': raw_id or '',
                'productName': raw_name or '',
                'numPurchases': cnt,
                'payAmount': amt,
            })

    if all_items and sales_list:
        logger.info(f"  📌 매핑: {matched_count}/{len(sales_list)}개 성공")
        if failed_samples:
            logger.info(f"  🔍 매핑 실패 샘플 (TOP 5):")
            for s in failed_samples:
                logger.info(s)

        # 매핑 실패 전체를 CSV로 저장 (수동 매핑용)
        if failed_all:
            try:
                import csv
                csv_path = os.path.join(os.getcwd(), 'unmapped_products.csv')
                # 기존 파일에 누적 (중복 제거를 위해 productId 기준)
                existing = {}
                if os.path.exists(csv_path):
                    with open(csv_path, 'r', encoding='utf-8-sig') as f:
                        reader = csv.DictReader(f)
                        for row in reader:
                            existing[row.get('productId', '')] = row

                # 새 데이터 합치기 (수량/매출은 누적)
                for f_item in failed_all:
                    pid = str(f_item['productId'])
                    if pid in existing:
                        existing[pid]['numPurchases'] = int(existing[pid].get('numPurchases', 0) or 0) + f_item['numPurchases']
                        existing[pid]['payAmount'] = int(existing[pid].get('payAmount', 0) or 0) + int(f_item['payAmount'])
                    else:
                        existing[pid] = {
                            'productId': pid,
                            'productName': f_item['productName'],
                            'numPurchases': f_item['numPurchases'],
                            'payAmount': int(f_item['payAmount']),
                        }

                # 매출 큰 순 정렬
                sorted_rows = sorted(existing.values(),
                                      key=lambda x: int(x.get('payAmount', 0) or 0),
                                      reverse=True)

                with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=['productId', 'productName', 'numPurchases', 'payAmount'])
                    writer.writeheader()
                    writer.writerows(sorted_rows)

                logger.info(f"  📋 매핑 실패 전체 ({len(failed_all)}개) → unmapped_products.csv 저장 (누적 {len(sorted_rows)}개)")
            except Exception as e:
                logger.warning(f"  ⚠️ unmapped CSV 저장 실패: {e}")

    # 원본 (raw) 상품 리스트 - 키워드 매칭 등 다른 곳에서 raw 이름이 필요할 때 사용
    stats["products_raw"] = sorted(sales_list, key=lambda x: x.get('payAmount', 0), reverse=True)

    # 베스트셀러용: 같은 _master_name으로 합치기
    merged = {}
    for p in sales_list:
        key = p.get('_master_name') or p.get('productName', '')
        if key not in merged:
            merged[key] = {
                '_master_name': p.get('_master_name'),
                'productName': p.get('productName', ''),
                '_category': p.get('_category', '미분류'),
                '_line': p.get('_line', '?'),
                '_unit_cost': p.get('_unit_cost', 0),
                'payAmount': 0,
                'numPurchases': 0,
                '_cogs': 0,
                '_gp': 0,
                '_margin_rate': 0,
                # productId는 머지 시 다 묶이므로 첫번째 것을 보관
                'productId': p.get('productId'),
            }
        merged[key]['payAmount'] += p.get('payAmount', 0) or 0
        merged[key]['numPurchases'] += int(p.get('numPurchases', 0) or 0)
        merged[key]['_cogs'] += p.get('_cogs', 0) or 0
        merged[key]['_gp'] += p.get('_gp', 0) or 0

    # 머지된 항목의 마진율 재계산
    for m in merged.values():
        m['_margin_rate'] = (m['_gp'] / m['payAmount'] * 100) if m['payAmount'] > 0 else 0

    stats["products"] = sorted(merged.values(), key=lambda x: x['payAmount'], reverse=True)
    stats["total_revenue"] = sum(p.get('payAmount', 0) for p in sales_list)
    stats["total_purchases"] = sum(int(p.get('numPurchases', 0) or 0) for p in sales_list)
    stats["total_cogs"] = total_cogs
    stats["total_gp"] = total_gp
    stats["margin_rate"] = (total_gp / stats["total_revenue"] * 100) if stats["total_revenue"] > 0 else 0

    # 매핑률 50% 미만이면 원가/마진 신뢰 불가 → 표시 안 함
    match_rate = matched_count / len(sales_list) if sales_list else 0
    stats["cost_reliable"] = match_rate >= 0.5
    stats["match_rate"] = match_rate

    # --- 마케팅 (채널/디바이스/광고비)
    mkt_rows = data.get('mkt_all', {}).get('rows', []) or []
    stats["total_cost"] = sum(ch.get("cost", 0) for ch in mkt_rows)

    channel_map = {}
    for ch in mkt_rows:
        name = ch.get('channelName', '기타')
        if name not in channel_map:
            channel_map[name] = {'name': name, 'cost': 0, 'interactions': 0,
                                  'purchases': 0, 'revenue': 0}
        channel_map[name]['cost'] += ch.get('cost', 0)
        channel_map[name]['interactions'] += int(ch.get('numInteractions', 0) or 0)
        channel_map[name]['purchases'] += int(ch.get('numPurchases', 0) or 0)
        channel_map[name]['revenue'] += ch.get('payAmount', 0)
    for ch in channel_map.values():
        ch['cvr'] = (ch['purchases'] / ch['interactions'] * 100) if ch['interactions'] > 0 else 0
        ch['roas'] = (ch['revenue'] / ch['cost'] * 100) if ch['cost'] > 0 else 0
    stats["channels"] = sorted(channel_map.values(), key=lambda x: x['revenue'], reverse=True)

    # --- 키워드
    kw_rows = data.get('keyword', {}).get('rows', []) or []
    for kw in kw_rows:
        inter = kw.get('numInteractions', 0)
        kw['cvr'] = (kw.get('numPurchases', 0) / inter * 100) if inter > 0 else 0
    stats["keywords"] = sorted(kw_rows, key=lambda x: x.get('numInteractions', 0), reverse=True)

    # --- 시간대별 + 환불
    hourly_raw = data.get('hourly_s', {}).get('rows', []) or []
    stats["total_refund"] = sum(h.get('refundPayAmount', 0) for h in hourly_raw)
    stats["total_cancel"] = sum(
        (h.get('cancelPayAmount', 0) or
         h.get('cancelAmount', 0) or
         h.get('canceledPayAmount', 0) or 0)
        for h in hourly_raw
    )
    # 환불 = 취소 + 반품 통합
    stats["total_returns"] = stats["total_refund"] + stats["total_cancel"]
    # 실매출 = 결제 - 환불
    stats["net_revenue"] = stats["total_revenue"] - stats["total_returns"]
    hour_map = {}
    for h in hourly_raw:
        hk = int(h.get('hour', 0))
        if hk not in hour_map:
            hour_map[hk] = {'payAmount': 0, 'purchases': 0}
        hour_map[hk]['payAmount'] += h.get('payAmount', 0)
        hour_map[hk]['purchases'] += int(h.get('numPurchases', 0) or 0)
    stats["hourly_sales"] = [{"hour": f"{h:02d}", **v} for h, v in sorted(hour_map.items())]

    # --- v2.0: 카테고리별 매출 (마스터 카탈로그 매핑 결과로 자동 계산)
    # 네이버 category_sales API 응답이 못 쓸만 해서, 우리가 매핑한 Type 기준으로 직접 계산
    cat_map = {}
    for p in stats["products"]:
        cat = p.get('_category', '미분류')
        amt = p.get('payAmount', 0) or 0
        pur = p.get('numPurchases', 0) or 0
        gp = p.get('_gp', 0) or 0
        if cat not in cat_map:
            cat_map[cat] = {'name': cat, 'revenue': 0, 'purchases': 0, 'gp': 0}
        cat_map[cat]['revenue'] += amt
        cat_map[cat]['purchases'] += int(pur)
        cat_map[cat]['gp'] += gp
    total_cat_rev = sum(c['revenue'] for c in cat_map.values()) or stats["total_revenue"] or 1
    for c in cat_map.values():
        c['share'] = c['revenue'] / total_cat_rev * 100
        c['margin_rate'] = (c['gp'] / c['revenue'] * 100) if c['revenue'] > 0 else 0
    stats["categories"] = sorted(cat_map.values(), key=lambda x: x['revenue'], reverse=True)

    # --- v2.0 신규: 라인별 매출 (시리즈별 — GO/V/EXP/RFIDsafe 등)
    line_map = {}
    for p in stats["products"]:
        line = p.get('_line', '?')
        if line == '?':
            continue  # 매핑 안 된 상품은 라인 분석에서 제외
        amt = p.get('payAmount', 0) or 0
        pur = p.get('numPurchases', 0) or 0
        gp = p.get('_gp', 0) or 0
        if line not in line_map:
            line_map[line] = {'name': line, 'revenue': 0, 'purchases': 0, 'gp': 0}
        line_map[line]['revenue'] += amt
        line_map[line]['purchases'] += int(pur)
        line_map[line]['gp'] += gp
    total_line_rev = sum(l['revenue'] for l in line_map.values()) or 1
    for l in line_map.values():
        l['share'] = l['revenue'] / total_line_rev * 100
        l['margin_rate'] = (l['gp'] / l['revenue'] * 100) if l['revenue'] > 0 else 0
    stats["lines"] = sorted(line_map.values(), key=lambda x: x['revenue'], reverse=True)

    # --- v2.0: 상품별 유입 키워드 (베스트셀러 기준, 마스터 이름 + 노이즈 제거)
    pk_raw = data.get('product_keyword', {})
    pk_rows = _extract_rows(pk_raw, 'rows', 'keywordByProductReport', 'productSearchReport')

    # 디버그: 응답 구조 한 번 출력
    if all_items and pk_rows and len(pk_rows) > 0 and isinstance(pk_rows[0], dict):
        first = pk_rows[0]
        logger.info(f"  🔍 [디버그] product_keyword 첫 row 키들: {list(first.keys())}")
        logger.info(f"     첫 row 샘플: productId={first.get('productId') or first.get('channelProductNo')}, keyword={first.get('refKeyword') or first.get('keyword')}, interactions={first.get('numInteractions') or first.get('numClicks')}")
    elif all_items and not pk_rows:
        logger.warning(f"  ⚠️ product_keyword API 응답이 비어있음. raw: {str(pk_raw)[:200]}")

    # 베스트셀러 productName → 마스터 이름 매핑
    # (네이버 product_keyword API는 productId 안 줘서 raw 이름으로 매칭)
    # products는 머지된 것이라 raw 이름 1개만 갖고 있어서 products_raw 사용
    bestseller_by_name = {}
    raw_products = stats.get("products_raw", stats["products"])
    # 머지된 베스트셀러 8개의 master_name 집합 만들고
    top_master_names = set()
    for idx, p in enumerate(stats["products"][:8]):
        mn = p.get('_master_name')
        if mn:
            top_master_names.add(mn)
    # raw 상품 중 top master에 속하는 것들의 raw 이름 → master 매핑
    for p in raw_products:
        master_name = p.get('_master_name')
        raw_name = p.get('productName', '')
        if master_name in top_master_names and raw_name:
            # 같은 마스터의 여러 raw 이름이 있을 수 있으니 모두 등록
            if raw_name not in bestseller_by_name:
                # order는 머지된 베스트셀러에서의 순위
                order = next((i for i, mp in enumerate(stats["products"][:8])
                              if mp.get('_master_name') == master_name), 0)
                bestseller_by_name[raw_name] = {
                    'master_name': master_name,
                    'order': order
                }

    # 노이즈 키워드 (검색 정보 없는 것들)
    NOISE_KEYWORDS = {'(notprovided)', '(not provided)', '(direct)', '(검색어 없음)', '-', '', None}

    pk_map = {}
    for r in pk_rows:
        if not isinstance(r, dict):
            continue

        # raw 상품명으로 베스트셀러 매칭
        prod_raw = r.get('productName', '')
        if prod_raw not in bestseller_by_name:
            continue
        master_info = bestseller_by_name[prod_raw]

        # 키워드 + 노이즈 필터
        kw = (r.get('refKeyword') or r.get('keyword') or
              r.get('searchKeyword') or '?')
        if kw in NOISE_KEYWORDS:
            continue

        # 메트릭 (이 API는 payAmount만 옴)
        revenue = r.get('payAmount', 0) or 0
        inter = (r.get('numInteractions') or r.get('numClicks') or
                 r.get('clicks') or r.get('interactions') or 0)
        purchases = (r.get('numPurchases') or r.get('purchases') or 0)

        master_name = master_info['master_name']
        if master_name not in pk_map:
            pk_map[master_name] = {'product': master_name, 'order': master_info['order'], 'keywords': []}
        pk_map[master_name]['keywords'].append({
            'keyword': kw,
            'interactions': int(inter or 0),
            'purchases': int(purchases or 0),
            'revenue': revenue,
        })

    # 키워드 정렬 (매출 순 우선, 조회수 차순), 상위 3개만
    for p in pk_map.values():
        p['keywords'] = sorted(p['keywords'],
                                key=lambda x: (x['revenue'], x['interactions']),
                                reverse=True)[:3]
    # 베스트셀러 순서대로
    stats["product_keywords"] = sorted(pk_map.values(), key=lambda x: x['order'])[:5]

    # --- v2.0: 고객 신규/기존
    cs_rows = _extract_rows(data.get('customer_status', {}), 'rows', 'customerStatusReport', 'statisticsReport')
    for r in cs_rows:
        if not isinstance(r, dict):
            continue
        new_cnt = r.get('newCustomerCount') or r.get('numNewCustomers') or 0
        ret_cnt = r.get('returningCustomerCount') or r.get('numReturningCustomers') or 0
        stats["customer_new"] += new_cnt
        stats["customer_returning"] += ret_cnt

    # --- v2.0: 재구매율
    rep_rows = _extract_rows(data.get('repurchase', {}), 'rows', 'repurchaseReport', 'statisticsReport')
    if rep_rows:
        latest = rep_rows[-1] if isinstance(rep_rows[-1], dict) else {}
        stats["repurchase_rate"] = (latest.get('repurchaseRate') or
                                     latest.get('rate') or
                                     latest.get('ratio') or 0)

    return stats


# =========================================================
# 3. 위험/기회 신호 자동 탐지
# =========================================================
def detect_alerts(curr, prev, prev_label="전일"):
    """매출/광고/상품 변화에서 위험·기회 신호 추출"""
    risks, opportunities = [], []

    # 채널 ROAS 하락 (임계값 -20% 이상, 광고비 5만원 이상)
    for ch in curr.get("channels", []):
        prev_ch = next((p for p in prev.get("channels", []) if p['name'] == ch['name']), None)
        if prev_ch and prev_ch.get('roas', 0) > 0 and ch.get('cost', 0) > 50000:
            change = pct_change(ch['roas'], prev_ch['roas'])
            if change is not None and change < -20:
                risks.append({
                    "title": f"{ch['name']} ROAS 하락",
                    "desc": f"{prev_label} {prev_ch['roas']:.0f}% → 현재 {ch['roas']:.0f}% ({change:+.0f}%p). 입찰가/소재 점검 필요."
                })

    # 채널 ROAS 100% 미만 (적자 채널)
    for ch in curr.get("channels", []):
        if ch.get('cost', 0) > 100000 and 0 < ch.get('roas', 0) < 100:
            risks.append({
                "title": f"{ch['name']} 손익분기 미달",
                "desc": f"광고비 {format_curr(ch['cost'])}, 매출 {format_curr(ch['revenue'])} (ROAS {ch['roas']:.0f}%). 효율 점검 필요."
            })

    # 반품 증가 (임계값 +20% 이상, 30만원 초과)
    refund_change = pct_change(curr.get('total_refund', 0), prev.get('total_refund', 0))
    if refund_change is not None and refund_change > 20 and curr.get('total_refund', 0) > 300000:
        risks.append({
            "title": "반품액 증가",
            "desc": f"{prev_label} 대비 {refund_change:+.0f}% ({format_curr(curr['total_refund'])}). 반품 사유 점검 필요."
        })

    # 상품 매출 급감 (임계값 -40%)
    for p in curr.get("products", [])[:10]:
        prev_p = next((x for x in prev.get("products", []) if x.get('productName') == p.get('productName')), None)
        if prev_p:
            change = pct_change(p.get('payAmount', 0), prev_p.get('payAmount', 0))
            if change is not None and change < -40 and prev_p.get('payAmount', 0) > 500000:
                name = p.get('_master_name') or p.get('productName', '')[:40]
                risks.append({
                    "title": f"{name} 매출 급감",
                    "desc": f"{prev_label} 대비 {change:.0f}%. 재고/노출 점검 필요."
                })

    # 키워드 노출 증가 (기회) - 임계값 30% 이상, 30회+, 노이즈 제외
    NOISE_KW = {'(검색어 없음)', '(notprovided)', '(direct)', '-', '', None}
    for kw in curr.get("keywords", [])[:15]:
        kw_text = kw.get('refKeyword', '')
        if kw_text in NOISE_KW:
            continue
        prev_kw = next((x for x in prev.get("keywords", []) if x.get('refKeyword') == kw_text), None)
        if prev_kw:
            change = pct_change(kw.get('numInteractions', 0), prev_kw.get('numInteractions', 0))
            if change is not None and change > 30 and kw.get('numInteractions', 0) > 30:
                opportunities.append({
                    "title": f'"{kw_text}" 검색량 증가',
                    "desc": f"노출 {change:+.0f}%. 광고비 추가 배정 검토 가치."
                })

    # 고효율 채널 (ROAS 500% 이상 + 광고비 일정 이상) - 기회
    for ch in curr.get("channels", []):
        if ch.get('cost', 0) > 100000 and ch.get('roas', 0) > 500:
            # 같은 채널이 위험에도 들어갔으면 스킵
            if not any(ch['name'] in r['title'] for r in risks):
                opportunities.append({
                    "title": f"{ch['name']} 고효율 ({ch['roas']:.0f}%)",
                    "desc": f"광고비 {format_curr(ch['cost'])} → 매출 {format_curr(ch['revenue'])}. 예산 추가 배정 검토."
                })
                break  # 고효율 채널은 1개만

    # 재구매율 상승 (기회) - 비활성화 상태일 때는 스킵
    if curr.get("repurchase_rate", 0) > 0 and curr.get("repurchase_rate", 0) > prev.get("repurchase_rate", 0) + 3:
        opportunities.append({
            "title": f"재구매율 {curr['repurchase_rate']:.0f}% 상승 추세",
            "desc": f"{prev_label} {prev.get('repurchase_rate', 0):.0f}% → +{curr['repurchase_rate'] - prev.get('repurchase_rate', 0):.0f}%p. 충성도 상승 신호."
        })

    return risks[:3], opportunities[:3]


# =========================================================
# 4. AI 분석 (Opus 4.7 + Extended Thinking)
# =========================================================
def get_ai_analysis(curr, prev, prefix, report_type, anthropic_key):
    """Opus 4.7로 한 줄 요약 + 액션 플랜 생성 (상품명 매핑은 코드에서 이미 처리됨)"""
    # 비교 라벨 자동 결정
    prev_label_map = {"일간": "전일", "주간": "전주", "월간": "전월"}
    prev_label = prev_label_map.get(report_type, "전일")
    try:
        client = anthropic.Anthropic(api_key=anthropic_key)
    except Exception as e:
        logger.error(f"Anthropic 클라이언트 실패: {e}")
        return _fallback_analysis(curr)

    rev_change = pct_change(curr['total_revenue'], prev['total_revenue']) or 0
    cost_change = pct_change(curr['total_cost'], prev['total_cost']) or 0
    # 실매출 기반 ROAS (취소·반품 차감)
    curr_net = curr['total_revenue'] - curr['total_refund'] - curr.get('total_cancel', 0)
    prev_net = prev['total_revenue'] - prev['total_refund'] - prev.get('total_cancel', 0)
    roas = (curr_net / curr['total_cost'] * 100) if curr['total_cost'] > 0 else 0
    prev_roas = (prev_net / prev['total_cost'] * 100) if prev['total_cost'] > 0 else 0
    refund_change = pct_change(curr['total_refund'], prev['total_refund']) or 0
    cancel_change = pct_change(curr.get('total_cancel', 0), prev.get('total_cancel', 0)) or 0

    # 매핑된 베스트셀러 (TOP 8) - 마스터 이름 + 카테고리
    bestsellers = [
        f"{p.get('_master_name', p.get('productName', '')[:40])} ({p.get('_category', '?')}) - {format_curr(p.get('payAmount', 0))}"
        for p in curr['products'][:8]
    ]

    # 모드별 예시 톤
    if report_type == "월간":
        period_subj = "지난달은"
        action_subj = "이번달 액션 플랜"
        oneliner_example = '"지난달은 매출 2억 5천만원으로 전월보다 8% 늘었고, ROAS도 320%로 올랐어요. 반품액이 18% 증가한 점이 눈에 띄네요."'
    elif report_type == "주간":
        period_subj = "지난주는"
        action_subj = "이번주 액션 플랜"
        oneliner_example = '"지난주는 매출 5,800만원으로 전주보다 5% 줄었지만, ROAS는 415%로 개선됐어요. 반품도 11% 감소했네요."'
    else:
        period_subj = "어제는"
        action_subj = "내일 액션 플랜"
        oneliner_example = '"어제는 매출 825만원으로 전일보다 12% 늘었고, ROAS도 348%로 올랐어요. 반품도 줄었네요."'

    prompt = f"""당신은 '{prefix}' 브랜드의 데이터 분석가입니다.
'{report_type}' 데이터를 분석하세요.

[실적 데이터]
- 결제 매출: {format_curr(curr['total_revenue'])} ({rev_change:+.1f}% vs {prev_label})
- 취소: {format_curr(curr.get('total_cancel', 0))} ({cancel_change:+.1f}% vs {prev_label})
- 반품: {format_curr(curr['total_refund'])} ({refund_change:+.1f}% vs {prev_label})
- 실매출: {format_curr(curr_net)} (= 결제 - 취소 - 반품)
- 광고비: {format_curr(curr['total_cost'])} ({cost_change:+.1f}% vs {prev_label})
- ROAS: {roas:.0f}% ({prev_label} {prev_roas:.0f}%, 실매출 기준)
- 판매량: {int(curr['total_purchases'])}건
{("- 매출총이익: " + format_curr(curr.get('total_gp', 0)) + " (마진율 " + f"{curr.get('margin_rate', 0):.0f}" + "%)") if _show_margin(curr) else ""}

[베스트셀러 TOP 8]
{chr(10).join(bestsellers)}

[작업]

1. one_liner (한 줄 요약)
   - 첫 단어는 "{period_subj}"로 시작
   - 핵심 숫자 2~3개를 자연어로 풀어서
   - 의인화/해석/과장 금지 ("분위기 좋은", "확장기 진입" 같은 표현 X)
   - 부정적 변화도 사실 그대로
   - 자연스러운 어미 (~네요, ~어요)
   - 예: {oneliner_example}

2. actions ({action_subj} 3가지)
   - 각각: title + desc(왜 그래야 하는지 1문장)
   - 데이터 근거 있는 구체적 액션
   - 베스트셀러 카테고리/시리즈 분포 활용
   - 예시: {{"title": "GO 시리즈 광고 강화", "desc": "GO 카테고리가 매출 60% 차지, 추가 투자 효율 높음"}}

[출력 포맷] - 반드시 JSON만, 다른 텍스트 없이
{{"one_liner": "...", "actions": [{{"title":"...", "desc":"..."}}, ...]}}
"""

    try:
        response = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}]
        )
        text = ""
        for block in response.content:
            if hasattr(block, 'text'):
                text = block.text

        # JSON 파싱
        text = text.strip()
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        text = text.strip()

        result = json.loads(text)

        # 안전장치
        if not result.get("one_liner"):
            result["one_liner"] = _fallback_one_liner(curr, prev)
        if not result.get("actions"):
            result["actions"] = []

        return result
    except Exception as e:
        logger.error(f"AI 분석 실패: {e}")
        return _fallback_analysis(curr, prev)


def _fallback_one_liner(curr, prev):
    rev_change = pct_change(curr['total_revenue'], prev['total_revenue']) or 0
    roas = (curr['total_revenue'] / curr['total_cost'] * 100) if curr['total_cost'] > 0 else 0
    return f"매출 {format_curr(curr['total_revenue'])}원 ({rev_change:+.1f}%), ROAS {roas:.0f}%."


def _fallback_analysis(curr, prev=None):
    return {
        "one_liner": _fallback_one_liner(curr, prev or {"total_revenue": 0, "total_cost": 0}),
        "actions": []
    }


# =========================================================
# 5. HTML 생성 (v2.0 디자인)
# =========================================================
def render_html(prefix, prefix_en, stats, prev_stats, ai, risks, opps, start, end, report_type):
    """v2.0 디자인 HTML 생성"""
    # 실매출 = 결제 - 취소 - 반품
    net_rev = stats["total_revenue"] - stats["total_refund"] - stats.get("total_cancel", 0)
    roas = (net_rev / stats["total_cost"] * 100) if stats["total_cost"] > 0 else 0
    prev_net = prev_stats["total_revenue"] - prev_stats["total_refund"] - prev_stats.get("total_cancel", 0)
    prev_roas = (prev_net / prev_stats["total_cost"] * 100) if prev_stats["total_cost"] > 0 else 0

    new_cust = stats.get("customer_new", 0)
    ret_cust = stats.get("customer_returning", 0)
    total_cust = (new_cust + ret_cust) or 1
    new_rate = new_cust / total_cust * 100
    repurchase = stats.get("repurchase_rate", 0)
    prev_repurchase = prev_stats.get("repurchase_rate", 0)

    dt_start = datetime.strptime(start, "%Y-%m-%d")
    dt_end = datetime.strptime(end, "%Y-%m-%d")
    weekday_kr = ["월", "화", "수", "목", "금", "토", "일"]

    # 보조 KPI 카드
    if _show_margin(stats):
        # 마진 가능 → 매출총이익 + 광고 후 이익
        cost_kpi_html = f"""<div style="display:grid; grid-template-columns:repeat(2,1fr); gap:8px; margin-top:12px;">
    <div style="background:linear-gradient(135deg,#fef9c3,#fef3c7); padding:12px; border-radius:8px; text-align:center; border:1px solid #fde047;">
      <div style="font-size:10px; color:#854d0e; margin-bottom:4px; font-weight:600;">💵 매출총이익</div>
      <div style="font-size:16px; font-weight:800;">{format_curr(stats.get('total_gp', 0))}</div>
      <div style="font-size:10px; color:#64748b; margin-top:2px;">마진율 {stats.get('margin_rate', 0):.0f}%</div>
    </div>
    <div style="background:linear-gradient(135deg,#dbeafe,#bfdbfe); padding:12px; border-radius:8px; text-align:center; border:1px solid #93c5fd;">
      <div style="font-size:10px; color:#1e40af; margin-bottom:4px; font-weight:600;">📊 광고 후 이익</div>
      <div style="font-size:16px; font-weight:800;">{format_curr(stats.get('total_gp', 0) - stats.get('total_cost', 0))}</div>
      <div style="font-size:10px; color:#64748b; margin-top:2px;">광고비 {format_curr(stats['total_cost'])} 차감</div>
    </div>
  </div>"""
    else:
        # 마진 없음 → 광고비 + 평균 객단가
        purchases = int(stats.get('total_purchases', 0) or 0)
        net_revenue = stats.get('total_revenue', 0) - stats.get('total_refund', 0) - stats.get('total_cancel', 0)
        avg_order = (net_revenue / purchases) if purchases > 0 else 0
        cost_kpi_html = f"""<div style="display:grid; grid-template-columns:repeat(2,1fr); gap:8px; margin-top:12px;">
    <div style="background:#f8fafc; padding:12px; border-radius:8px; text-align:center; border:1px solid #e2e8f0;">
      <div style="font-size:10px; color:#64748b; margin-bottom:4px; font-weight:600;">📢 광고비</div>
      <div style="font-size:16px; font-weight:800;">{format_curr(stats.get('total_cost', 0))}</div>
      <div style="font-size:10px; color:#64748b; margin-top:2px;">판매 {purchases}건</div>
    </div>
    <div style="background:#f8fafc; padding:12px; border-radius:8px; text-align:center; border:1px solid #e2e8f0;">
      <div style="font-size:10px; color:#64748b; margin-bottom:4px; font-weight:600;">🛒 평균 객단가</div>
      <div style="font-size:16px; font-weight:800;">{format_curr(avg_order)}</div>
      <div style="font-size:10px; color:#64748b; margin-top:2px;">실매출 ÷ 판매건수</div>
    </div>
  </div>"""

    if report_type == "월간":
        date_label = f"{dt_start.year}년 {dt_start.month}월"
        oneliner_label = "지난달 한 줄"
        action_period = "TODO 이번달"
    elif report_type == "주간" or start != end:
        date_label = f"{dt_start.month}/{dt_start.day} ~ {dt_end.month}/{dt_end.day}"
        oneliner_label = "지난주 한 줄"
        action_period = "TODO 이번주"
    else:
        date_label = f"{dt_start.month}월 {dt_start.day}일 ({weekday_kr[dt_start.weekday()]})"
        oneliner_label = "어제의 한 줄"
        action_period = "TODO 내일"

    # 베스트셀러 행 (TOP 8, 마진 제거 + "기타" 합계 행)
    bestseller_html = ""
    for i, p in enumerate(stats["products"][:8]):
        name = p.get('_master_name') or p.get('productName', '')[:60]
        cat = p.get('_category', '미분류')
        amt = p.get('payAmount', 0)
        cnt = int(p.get('numPurchases', 0) or 0)
        # 원가 신뢰성 있을 때만 마진 표시
        margin_info = ""
        if _show_margin(stats):
            margin = p.get('_margin_rate', 0)
            if margin > 0:
                m_color = "#10b981" if margin >= 50 else ("#3b82f6" if margin >= 30 else "#64748b")
                margin_info = f" · 마진 <span style='color:{m_color}; font-weight:700;'>{margin:.0f}%</span>"
        bg = "background: linear-gradient(90deg, #fef3c7 0%, transparent 60%);" if i == 0 else ""
        rank_bg = ("linear-gradient(135deg, #fbbf24 0%, #f59e0b 100%)" if i == 0 else "#e2e8f0")
        rank_color = "white" if i == 0 else "#64748b"
        bestseller_html += f"""
        <div style="display:flex; align-items:center; gap:12px; padding:12px; {bg} border-radius:10px;">
          <div style="width:32px; height:32px; background:{rank_bg}; color:{rank_color}; border-radius:8px; display:flex; align-items:center; justify-content:center; font-weight:800; font-size:14px;">{i+1}</div>
          <div style="flex:1; min-width:0;">
            <div style="font-weight:600; font-size:13px; line-height:1.4;">{name}</div>
            <div style="font-size:11px; color:#64748b; margin-top:2px;">{cat} · {cnt}개 판매{margin_info}</div>
          </div>
          <div style="font-weight:800; font-size:15px; white-space:nowrap;">{format_curr(amt)}</div>
        </div>"""

    # "기타" 행: 9위 이하 상품 합계
    others = stats["products"][8:]
    if others:
        other_amt = sum(p.get('payAmount', 0) for p in others)
        other_cnt = sum(int(p.get('numPurchases', 0) or 0) for p in others)
        bestseller_html += f"""
        <div style="display:flex; align-items:center; gap:12px; padding:12px; background:#f8fafc; border-radius:10px; margin-top:4px; border:1px dashed #cbd5e1;">
          <div style="width:32px; height:32px; background:#cbd5e1; color:#475569; border-radius:8px; display:flex; align-items:center; justify-content:center; font-weight:800; font-size:11px;">기타</div>
          <div style="flex:1; min-width:0;">
            <div style="font-weight:600; font-size:13px; line-height:1.4; color:#475569;">기타 {len(others)}개 상품</div>
            <div style="font-size:11px; color:#64748b; margin-top:2px;">{other_cnt}개 판매</div>
          </div>
          <div style="font-weight:800; font-size:15px; white-space:nowrap; color:#475569;">{format_curr(other_amt)}</div>
        </div>"""

    # 카테고리 막대 (analyze_data에서 이미 매핑 기반으로 계산됨)
    cat_colors = [("#3b82f6", "#60a5fa"), ("#10b981", "#34d399"),
                  ("#f59e0b", "#fbbf24"), ("#8b5cf6", "#a78bfa"),
                  ("#ef4444", "#f87171"), ("#06b6d4", "#22d3ee"),
                  ("#ec4899", "#f472b6"), ("#14b8a6", "#5eead4")]
    cat_html = ""
    for i, c in enumerate(stats.get("categories", [])[:6]):
        c1, c2 = cat_colors[i % len(cat_colors)]
        share = c.get('share', 0)
        margin = c.get('margin_rate', 0)
        margin_badge = f"<span style='font-size:10px; color:#64748b; font-weight:500; margin-left:6px;'>마진 {margin:.0f}%</span>" if (margin > 0 and _show_margin(stats)) else ""
        cat_html += f"""
        <div style="margin-bottom:12px;">
          <div style="display:flex; justify-content:space-between; margin-bottom:4px;">
            <span style="font-size:13px; font-weight:600;">{c['name']}{margin_badge}</span>
            <span style="font-size:13px; font-weight:700; color:{c1};">{share:.0f}% · {format_curr(c['revenue'])}</span>
          </div>
          <div style="height:10px; background:#f1f5f9; border-radius:5px; overflow:hidden;">
            <div style="height:100%; width:{share:.1f}%; background:linear-gradient(90deg, {c1} 0%, {c2} 100%); border-radius:5px;"></div>
          </div>
        </div>"""
    if not cat_html:
        cat_html = "<div style='color:#94a3b8; font-size:13px; padding:20px; text-align:center;'>카테고리 데이터 없음</div>"

    # v2.0 신규: 라인별 매출 (시리즈별)
    line_html = ""
    for i, l in enumerate(stats.get("lines", [])[:6]):
        c1, c2 = cat_colors[i % len(cat_colors)]
        share = l.get('share', 0)
        margin = l.get('margin_rate', 0)
        margin_badge = f"<span style='font-size:10px; color:#64748b; font-weight:500; margin-left:6px;'>마진 {margin:.0f}%</span>" if (margin > 0 and _show_margin(stats)) else ""
        line_html += f"""
        <div style="margin-bottom:12px;">
          <div style="display:flex; justify-content:space-between; margin-bottom:4px;">
            <span style="font-size:13px; font-weight:600;">{l['name']}{margin_badge}</span>
            <span style="font-size:13px; font-weight:700; color:{c1};">{share:.0f}% · {format_curr(l['revenue'])}</span>
          </div>
          <div style="height:10px; background:#f1f5f9; border-radius:5px; overflow:hidden;">
            <div style="height:100%; width:{share:.1f}%; background:linear-gradient(90deg, {c1} 0%, {c2} 100%); border-radius:5px;"></div>
          </div>
        </div>"""
    if not line_html:
        line_html = "<div style='color:#94a3b8; font-size:13px; padding:20px; text-align:center;'>라인 데이터 없음</div>"

    # 채널 행
    channel_html = ""
    for ch in stats["channels"][:8]:
        cvr_color = "#10b981" if ch.get('cvr', 0) >= 3 else "#64748b"
        channel_html += f"""
        <tr style="border-bottom:1px solid #f1f5f9;">
          <td style="padding:12px 6px; font-weight:600;">{ch['name']}</td>
          <td style="text-align:right; padding:12px 6px;">{int(ch['interactions'] or 0):,}</td>
          <td style="text-align:right; padding:12px 6px; color:{cvr_color}; font-weight:700;">{ch['cvr']:.1f}%</td>
          <td style="text-align:right; padding:12px 6px; font-weight:800;">{format_curr(ch['revenue'])}</td>
        </tr>"""

    # 키워드 태그 (노이즈 키워드 제외, 정수 표시)
    NOISE_KEYWORDS = {'(검색어 없음)', '(notprovided)', '(not provided)', '(direct)', '-', '', None}
    kw_html = ""
    kw_count = 0
    for kw in stats["keywords"]:
        kw_text = kw.get('refKeyword', '')
        if kw_text in NOISE_KEYWORDS:
            continue
        inter = int(kw.get('numInteractions', 0) or 0)
        kw_html += f"""<div style="background:#f1f5f9; color:#475569; padding:8px 14px; border-radius:16px; font-size:13px; font-weight:600;">{kw_text} <span style="opacity:0.6; font-size:11px;">{inter:,}</span></div>"""
        kw_count += 1
        if kw_count >= 10:
            break

    # v2.0 신규: 시간대별 매출 (일간만 의미 있음)
    hourly_section_html = ""
    if report_type == "일간":
        hourly_data = stats.get("hourly_sales", [])
        hourly_html = ""
        peak_hour = "-"
        peak_amt_text = ""
        if hourly_data:
            amts = {int(h.get('hour', 0)): h.get('payAmount', 0) for h in hourly_data}
            full_hours = [(i, amts.get(i, 0)) for i in range(24)]
            max_amt = max([a for _, a in full_hours]) or 1
            peak = max(full_hours, key=lambda x: x[1])
            peak_hour = f"{peak[0]}시"
            peak_amt_text = format_curr(peak[1])
            sorted_hours = sorted(full_hours, key=lambda x: x[1], reverse=True)
            top3_hours = {h for h, _ in sorted_hours[:3] if _ > 0}

            for h, amt in full_hours:
                ratio = amt / max_amt * 100
                if ratio < 1:
                    color1, color2 = "#e2e8f0", "#f1f5f9"
                elif ratio < 30:
                    color1, color2 = "#bfdbfe", "#dbeafe"
                elif ratio < 60:
                    color1, color2 = "#60a5fa", "#93c5fd"
                elif ratio < 90:
                    color1, color2 = "#3b82f6", "#60a5fa"
                else:
                    color1, color2 = "#1d4ed8", "#2563eb"
                label_html = ""
                if h in top3_hours:
                    label_html = f"""<div style="position:absolute; bottom:100%; left:50%; transform:translateX(-50%); font-size:9px; font-weight:700; color:#1e40af; white-space:nowrap; padding-bottom:2px;">{format_curr(amt)}</div>"""
                hourly_html += f"""<div style="flex:1; position:relative; display:flex; flex-direction:column; justify-content:flex-end; height:100%;">{label_html}<div style="height:{max(ratio, 1):.0f}%; background:linear-gradient(180deg,{color1},{color2}); border-radius:3px 3px 0 0;" title="{h}시 {format_curr(amt)}"></div></div>"""
        if not hourly_html:
            hourly_html = "<div style='color:#94a3b8; font-size:13px; padding:20px; text-align:center;'>시간대별 데이터 없음</div>"

        hourly_section_html = f"""<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">⏰ HOURLY SALES</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:8px;">시간대별 매출</h2>
  <div style="font-size:12px; color:#64748b; margin-bottom:24px;">피크 시간 <strong style="color:#3b82f6;">{peak_hour}</strong> · {peak_amt_text}</div>
  <div style="display:flex; align-items:flex-end; gap:3px; height:120px; padding:14px 4px 0 4px;">{hourly_html}</div>
  <div style="display:flex; justify-content:space-between; margin-top:6px; padding:0 4px; font-size:9px; color:#94a3b8;">
    <span>0시</span><span>6시</span><span>12시</span><span>18시</span><span>23시</span>
  </div>
</div>"""

    # ═════════════════════════════════════════════════════════════
    # 주간 전용 섹션
    # ═════════════════════════════════════════════════════════════
    weekly_sections_html = ""
    if report_type == "주간":
        weekday_kr = ["월", "화", "수", "목", "금", "토", "일"]
        breakdown = stats.get('daily_breakdown', [])

        # 1. 요일별 매출 패턴
        if breakdown:
            max_rev = max(d['revenue'] for d in breakdown) or 1
            day_bars = ""
            weekday_totals = {i: {'revenue': 0, 'count': 0} for i in range(7)}
            for d in breakdown:
                wd = d['date'].weekday()
                weekday_totals[wd]['revenue'] += d['revenue']
                weekday_totals[wd]['count'] += 1

            avg_weekday = []
            for wd in range(7):
                wt = weekday_totals[wd]
                avg = wt['revenue'] / wt['count'] if wt['count'] > 0 else 0
                avg_weekday.append((wd, avg))
            max_avg = max(a for _, a in avg_weekday) or 1

            wd_bars = ""
            for wd, avg in avg_weekday:
                ratio = avg / max_avg * 100 if max_avg > 0 else 0
                is_weekend = wd >= 5
                bar_color = "#f59e0b" if is_weekend else "#3b82f6"
                wd_bars += f"""
                <div style="display:flex; align-items:center; gap:10px; margin-bottom:6px;">
                  <div style="width:24px; font-size:12px; font-weight:700; color:{'#92400e' if is_weekend else '#1e40af'};">{weekday_kr[wd]}</div>
                  <div style="flex:1; height:18px; background:#f1f5f9; border-radius:4px; position:relative; overflow:hidden;">
                    <div style="height:100%; width:{ratio:.1f}%; background:{bar_color}; border-radius:4px;"></div>
                  </div>
                  <div style="width:80px; text-align:right; font-size:12px; font-weight:700;">{format_curr(avg)}</div>
                </div>"""

            # 평일/주말 비교
            weekday_avg = sum(a for wd, a in avg_weekday[:5]) / 5 if avg_weekday else 0
            weekend_avg = sum(a for wd, a in avg_weekday[5:]) / 2 if avg_weekday else 0
            weekend_diff = pct_change(weekend_avg, weekday_avg) or 0

            weekly_sections_html += f"""
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">📅 WEEKDAY PATTERN</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:8px;">요일별 매출 패턴</h2>
  <div style="font-size:12px; color:#64748b; margin-bottom:18px;">평일 평균 대비 주말 <strong style="color:{'#10b981' if weekend_diff > 0 else '#ef4444'}">{weekend_diff:+.0f}%</strong></div>
  {wd_bars}
</div>"""

        # 2. 신규 진입 / 순위 급상승 상품
        rank_changes = []
        for p in stats.get('products', [])[:10]:
            rc = p.get('_rank_change')
            if rc is None:
                rank_changes.append({'type': 'new', 'name': p.get('_master_name') or p.get('productName', '')[:40], 'curr': stats['products'].index(p) + 1})
            elif rc >= 3:
                rank_changes.append({'type': 'up', 'name': p.get('_master_name') or p.get('productName', '')[:40], 'curr': stats['products'].index(p) + 1, 'prev': p.get('_prev_rank'), 'change': rc})

        if rank_changes:
            mover_html = ""
            for r in rank_changes[:5]:
                if r['type'] == 'new':
                    badge = '<span style="background:#dcfce7; color:#15803d; padding:2px 8px; border-radius:6px; font-size:10px; font-weight:700;">NEW</span>'
                    detail = f"<span style='color:#64748b;'>이번주 #{r['curr']} 진입</span>"
                else:
                    badge = f'<span style="background:#dbeafe; color:#1e40af; padding:2px 8px; border-radius:6px; font-size:10px; font-weight:700;">▲ {r["change"]}↑</span>'
                    detail = f"<span style='color:#64748b;'>#{r['prev']} → #{r['curr']}</span>"
                mover_html += f"""
                <div style="display:flex; justify-content:space-between; align-items:center; padding:10px 0; border-bottom:1px solid #f1f5f9;">
                  <div style="flex:1; font-size:13px; font-weight:600;">{r['name']}</div>
                  <div style="display:flex; align-items:center; gap:8px; font-size:11px;">{detail} {badge}</div>
                </div>"""

            weekly_sections_html += f"""
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">🚀 RISING STARS</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:8px;">신규 진입·급상승 상품</h2>
  <div style="font-size:12px; color:#64748b; margin-bottom:16px;">전주 대비 순위 급상승</div>
  {mover_html}
</div>"""

        # 3. 광고 효율 일별 추이 (요일별 ROAS)
        if breakdown:
            daily_roas_html = ""
            max_daily_roas = max((d['roas'] for d in breakdown), default=1) or 1
            for d in breakdown:
                wd = weekday_kr[d['date'].weekday()]
                ratio = d['roas'] / max_daily_roas * 100 if max_daily_roas > 0 else 0
                roas_color = "#10b981" if d['roas'] >= 300 else ("#f59e0b" if d['roas'] >= 100 else "#ef4444")
                daily_roas_html += f"""
                <div style="display:flex; align-items:center; gap:10px; margin-bottom:6px;">
                  <div style="width:60px; font-size:11px; color:#64748b;">{d['date'].month}/{d['date'].day} ({wd})</div>
                  <div style="flex:1; height:14px; background:#f1f5f9; border-radius:3px; overflow:hidden;">
                    <div style="height:100%; width:{ratio:.1f}%; background:{roas_color}; border-radius:3px;"></div>
                  </div>
                  <div style="width:60px; text-align:right; font-size:12px; font-weight:700; color:{roas_color};">{d['roas']:.0f}%</div>
                </div>"""

            weekly_sections_html += f"""
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">📈 DAILY ROAS</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:16px;">일별 광고 효율 추이</h2>
  {daily_roas_html}
</div>"""

        # 4. 재고 회전 경고 (판매 속도 기반)
        velocity_html = ""
        for p in stats.get('products', [])[:10]:
            cnt = int(p.get('numPurchases', 0) or 0)
            if cnt < 5:
                continue
            avg_per_day = cnt / 7  # 주간이니 7일 평균
            name = p.get('_master_name') or p.get('productName', '')[:40]
            velocity_html += f"""
            <div style="display:flex; justify-content:space-between; align-items:center; padding:10px 0; border-bottom:1px solid #f1f5f9;">
              <div style="flex:1; font-size:13px; font-weight:600;">{name}</div>
              <div style="font-size:12px; color:#64748b;">하루 평균 <strong style='color:#0f172a;'>{avg_per_day:.1f}개</strong></div>
            </div>"""

        if velocity_html:
            weekly_sections_html += f"""
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">⚡ SALES VELOCITY</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:8px;">판매 속도</h2>
  <div style="font-size:12px; color:#64748b; margin-bottom:16px;">하루 평균 판매량 (재고 회전 참고용)</div>
  {velocity_html}
</div>"""

    # ═════════════════════════════════════════════════════════════
    # 월간 전용 섹션
    # ═════════════════════════════════════════════════════════════
    monthly_sections_html = ""
    if report_type == "월간":
        breakdown = stats.get('daily_breakdown', [])
        weekday_kr = ["월", "화", "수", "목", "금", "토", "일"]

        # 1. 일별 매출 트렌드 차트
        if breakdown:
            max_rev = max(d['revenue'] for d in breakdown) or 1
            avg_rev = sum(d['revenue'] for d in breakdown) / len(breakdown) if breakdown else 0
            avg_ratio = avg_rev / max_rev * 100

            day_bars_html = ""
            for d in breakdown:
                ratio = d['revenue'] / max_rev * 100 if max_rev > 0 else 0
                is_weekend = d['date'].weekday() >= 5
                bar_color = "#f59e0b" if is_weekend else "#3b82f6"
                day_bars_html += f"""<div style="flex:1; min-width:0; position:relative; display:flex; align-items:flex-end; height:100%;"><div style="width:100%; height:{max(ratio, 1):.0f}%; background:{bar_color}; border-radius:2px 2px 0 0;" title="{d['date'].strftime('%m/%d')} {format_curr(d['revenue'])}"></div></div>"""

            monthly_sections_html += f"""
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">📈 DAILY TREND</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:8px;">일별 매출 흐름</h2>
  <div style="font-size:12px; color:#64748b; margin-bottom:18px;">일평균 매출 <strong style="color:#3b82f6;">{format_curr(avg_rev)}</strong> · 평일 🟦 / 주말 🟧</div>
  <div style="position:relative; height:140px; padding:0 4px;">
    <div style="position:absolute; left:0; right:0; top:{100-avg_ratio:.0f}%; border-top:1px dashed #94a3b8; opacity:0.6;"><span style="position:absolute; right:0; top:-9px; background:#f8fafc; padding:0 4px; font-size:9px; color:#64748b;">평균</span></div>
    <div style="display:flex; align-items:flex-end; gap:2px; height:100%;">{day_bars_html}</div>
  </div>
  <div style="display:flex; justify-content:space-between; margin-top:6px; padding:0 4px; font-size:9px; color:#94a3b8;">
    <span>{breakdown[0]['date'].strftime('%m/%d')}</span><span>{breakdown[len(breakdown)//2]['date'].strftime('%m/%d')}</span><span>{breakdown[-1]['date'].strftime('%m/%d')}</span>
  </div>
</div>"""

        # 2. BEST / WORST 일
        if breakdown:
            best_day = max(breakdown, key=lambda x: x['revenue'])
            worst_day = min(breakdown, key=lambda x: x['revenue'])
            best_wd = weekday_kr[best_day['date'].weekday()]
            worst_wd = weekday_kr[worst_day['date'].weekday()]

            monthly_sections_html += f"""
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">🏅 BEST / WORST DAY</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:16px;">최고·최저 매출일</h2>
  <div style="display:grid; grid-template-columns:1fr 1fr; gap:12px;">
    <div style="background:linear-gradient(135deg,#f0fdf4,#dcfce7); border:1px solid #86efac; padding:16px; border-radius:12px;">
      <div style="font-size:11px; color:#15803d; font-weight:700; margin-bottom:6px;">🌟 BEST</div>
      <div style="font-size:15px; font-weight:700;">{best_day['date'].month}월 {best_day['date'].day}일 ({best_wd})</div>
      <div style="font-size:22px; font-weight:800; color:#15803d; margin-top:4px;">{format_curr(best_day['revenue'])}</div>
      <div style="font-size:11px; color:#64748b; margin-top:6px;">ROAS {best_day['roas']:.0f}% · {int(best_day['purchases'])}건</div>
      <div style="font-size:11px; color:#64748b; margin-top:4px;">TOP: {best_day['top_product'][:30]}</div>
    </div>
    <div style="background:linear-gradient(135deg,#fef2f2,#fee2e2); border:1px solid #fca5a5; padding:16px; border-radius:12px;">
      <div style="font-size:11px; color:#991b1b; font-weight:700; margin-bottom:6px;">⚠️ WORST</div>
      <div style="font-size:15px; font-weight:700;">{worst_day['date'].month}월 {worst_day['date'].day}일 ({worst_wd})</div>
      <div style="font-size:22px; font-weight:800; color:#991b1b; margin-top:4px;">{format_curr(worst_day['revenue'])}</div>
      <div style="font-size:11px; color:#64748b; margin-top:6px;">ROAS {worst_day['roas']:.0f}% · {int(worst_day['purchases'])}건</div>
      <div style="font-size:11px; color:#64748b; margin-top:4px;">TOP: {worst_day['top_product'][:30]}</div>
    </div>
  </div>
</div>"""

        # 3. 신규 vs 단골 매출 비중
        new_cust = stats.get('customer_new', 0)
        ret_cust = stats.get('customer_returning', 0)
        if new_cust + ret_cust > 0:
            total_cust = new_cust + ret_cust
            new_pct = new_cust / total_cust * 100
            ret_pct = ret_cust / total_cust * 100
            monthly_sections_html += f"""
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">👥 CUSTOMER MIX</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:16px;">신규 vs 단골 비중</h2>
  <div style="display:flex; height:36px; border-radius:8px; overflow:hidden; margin-bottom:12px;">
    <div style="width:{new_pct:.1f}%; background:linear-gradient(90deg,#fbbf24,#f59e0b); display:flex; align-items:center; justify-content:center; color:white; font-weight:700; font-size:13px;">{new_pct:.0f}%</div>
    <div style="width:{ret_pct:.1f}%; background:linear-gradient(90deg,#8b5cf6,#7c3aed); display:flex; align-items:center; justify-content:center; color:white; font-weight:700; font-size:13px;">{ret_pct:.0f}%</div>
  </div>
  <div style="display:flex; justify-content:space-around; font-size:13px;">
    <div><span style="color:#f59e0b; font-weight:700;">●</span> 신규 <strong>{new_cust:,}명</strong></div>
    <div><span style="color:#7c3aed; font-weight:700;">●</span> 단골 <strong>{ret_cust:,}명</strong></div>
  </div>
</div>"""

        # 4. 카테고리/시리즈 순위 변동
        cat_changes_html = ""
        for c in stats.get('categories', [])[:6]:
            rc = c.get('rank_change')
            if rc is None:
                badge = '<span style="background:#f3f4f6; color:#6b7280; padding:2px 6px; border-radius:4px; font-size:10px;">NEW</span>'
            elif rc > 0:
                badge = f'<span style="background:#dcfce7; color:#15803d; padding:2px 6px; border-radius:4px; font-size:10px;">▲ {rc}↑</span>'
            elif rc < 0:
                badge = f'<span style="background:#fee2e2; color:#991b1b; padding:2px 6px; border-radius:4px; font-size:10px;">▼ {abs(rc)}↓</span>'
            else:
                badge = '<span style="color:#94a3b8; font-size:10px;">-</span>'
            cat_changes_html += f"""
            <div style="display:flex; justify-content:space-between; align-items:center; padding:8px 0; border-bottom:1px solid #f1f5f9;">
              <div style="font-size:13px; font-weight:600;">{c['name']}</div>
              <div style="display:flex; gap:8px; align-items:center;">
                <span style="font-size:12px; color:#64748b;">{c['share']:.0f}%</span> {badge}
              </div>
            </div>"""

        line_changes_html = ""
        for l in stats.get('lines', [])[:6]:
            rc = l.get('rank_change')
            if rc is None:
                badge = '<span style="background:#f3f4f6; color:#6b7280; padding:2px 6px; border-radius:4px; font-size:10px;">NEW</span>'
            elif rc > 0:
                badge = f'<span style="background:#dcfce7; color:#15803d; padding:2px 6px; border-radius:4px; font-size:10px;">▲ {rc}↑</span>'
            elif rc < 0:
                badge = f'<span style="background:#fee2e2; color:#991b1b; padding:2px 6px; border-radius:4px; font-size:10px;">▼ {abs(rc)}↓</span>'
            else:
                badge = '<span style="color:#94a3b8; font-size:10px;">-</span>'
            line_changes_html += f"""
            <div style="display:flex; justify-content:space-between; align-items:center; padding:8px 0; border-bottom:1px solid #f1f5f9;">
              <div style="font-size:13px; font-weight:600;">{l['name']}</div>
              <div style="display:flex; gap:8px; align-items:center;">
                <span style="font-size:12px; color:#64748b;">{l['share']:.0f}%</span> {badge}
              </div>
            </div>"""

        monthly_sections_html += f"""
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">🔄 RANK CHANGES</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:16px;">전월 대비 순위 변동</h2>
  <div style="display:grid; grid-template-columns:1fr 1fr; gap:20px;">
    <div>
      <div style="font-size:12px; font-weight:700; color:#64748b; margin-bottom:8px;">📦 카테고리</div>
      {cat_changes_html}
    </div>
    <div>
      <div style="font-size:12px; font-weight:700; color:#64748b; margin-bottom:8px;">🏷️ 시리즈</div>
      {line_changes_html}
    </div>
  </div>
</div>"""

        # 5. 키워드 변화
        kw_changes_html = ""
        for k in stats.get('keywords', [])[:8]:
            kw_text = k.get('refKeyword', '')
            if kw_text in {'(검색어 없음)', '(notprovided)', '-', None, ''}:
                continue
            rc = k.get('rank_change')
            if rc is None:
                badge = '<span style="background:#dcfce7; color:#15803d; padding:2px 8px; border-radius:6px; font-size:10px; font-weight:700;">NEW</span>'
            elif rc >= 3:
                badge = f'<span style="background:#dbeafe; color:#1e40af; padding:2px 8px; border-radius:6px; font-size:10px; font-weight:700;">▲ {rc}↑</span>'
            elif rc <= -3:
                badge = f'<span style="background:#fee2e2; color:#991b1b; padding:2px 8px; border-radius:6px; font-size:10px; font-weight:700;">▼ {abs(rc)}↓</span>'
            else:
                continue  # 변화 작으면 스킵
            kw_changes_html += f"""<div style="display:inline-flex; align-items:center; gap:6px; background:#f1f5f9; padding:6px 12px; border-radius:16px; margin:0 4px 8px 0;"><span style="font-size:12px; font-weight:600;">{kw_text}</span>{badge}</div>"""

        if kw_changes_html:
            monthly_sections_html += f"""
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">🔍 KEYWORD TRENDS</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:8px;">키워드 트렌드 변화</h2>
  <div style="font-size:12px; color:#64748b; margin-bottom:16px;">전월 대비 검색 키워드 순위 변동</div>
  <div>{kw_changes_html}</div>
</div>"""

        # 6. 누적 + 연 환산
        annual_estimate = stats['total_revenue'] * 12
        annual_gp_text = f"매출총이익 {format_curr(stats.get('total_gp', 0) * 12)} (마진율 {stats.get('margin_rate', 0):.0f}%)" if _show_margin(stats) else f"판매 {int(stats['total_purchases']) * 12:,}건 (월 평균)"
        monthly_sections_html += f"""
<div class="card" style="padding:24px; background:linear-gradient(135deg,#1e293b 0%,#0f172a 100%); color:white;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#94a3b8;">📊 ANNUAL PROJECTION</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:16px;">연 환산 추정</h2>
  <div style="font-size:13px; color:#cbd5e1; margin-bottom:8px;">이번달 페이스를 12개월 유지하면</div>
  <div style="font-size:32px; font-weight:900; color:#60a5fa;">{format_curr(annual_estimate)}원</div>
  <div style="font-size:12px; color:#94a3b8; margin-top:8px;">연 매출 추정 · {annual_gp_text}</div>
</div>"""

    # v2.0 신규: 상품별 유입 키워드 매트릭스
    pk_html = ""
    for i, pk in enumerate(stats.get("product_keywords", [])[:5]):
        rank_bg = "#fbbf24" if i == 0 else "#e2e8f0"
        rank_color = "white" if i == 0 else "#64748b"
        # 의미 있는 키워드만 (interactions > 0 또는 revenue > 0)
        meaningful = [kw for kw in pk.get("keywords", []) if kw.get('interactions', 0) > 0 or kw.get('revenue', 0) > 0]
        keywords_html = ""
        for kw in meaningful[:3]:
            # 메트릭 표시: 매출 있으면 매출, 없으면 클릭수
            if kw.get('revenue', 0) > 0:
                metric = format_curr(kw['revenue'])
            elif kw.get('purchases', 0) > 0:
                metric = f"{kw['purchases']}건"
            else:
                metric = f"{kw['interactions']:,}회"
            keywords_html += f"""<span style="background:#eff6ff; color:#1e40af; padding:4px 10px; border-radius:8px; font-size:12px; font-weight:600;">{kw['keyword']} <span style="opacity:0.6;">({metric})</span></span>"""
        if not keywords_html:
            keywords_html = "<span style='color:#94a3b8; font-size:12px;'>유의미한 검색 유입 없음</span>"
        border = "border-bottom:1px solid #f1f5f9;" if i < len(stats["product_keywords"][:5]) - 1 else ""
        pk_html += f"""
        <div style="padding:14px 0; {border}">
          <div style="font-size:13px; font-weight:700; color:#0f172a; margin-bottom:8px;">
            <span style="display:inline-block; width:20px; height:20px; background:{rank_bg}; color:{rank_color}; border-radius:4px; text-align:center; font-size:11px; line-height:20px; margin-right:6px;">{i+1}</span>
            {pk['product'][:50]}
          </div>
          <div style="display:flex; flex-wrap:wrap; gap:6px; padding-left:26px;">{keywords_html}</div>
        </div>"""
    if not pk_html:
        pk_html = "<div style='color:#94a3b8; font-size:13px; padding:20px; text-align:center;'>상품별 키워드 데이터 없음</div>"

    # 위험 신호
    risks_html = ""
    for r in risks:
        risks_html += f"""
        <div style="background:#fef2f2; border-left:3px solid #ef4444; padding:14px 16px; border-radius:8px; margin-bottom:8px;">
          <div style="font-weight:600; font-size:14px; color:#991b1b;">{r['title']}</div>
          <div style="font-size:12px; color:#64748b; margin-top:4px;">{r['desc']}</div>
        </div>"""
    if not risks:
        risks_html = "<div style='font-size:13px; color:#94a3b8; padding:8px 0;'>특별한 위험 신호 없음 ✓</div>"

    # 기회 신호
    opps_html = ""
    for o in opps:
        opps_html += f"""
        <div style="background:#f0fdf4; border-left:3px solid #10b981; padding:14px 16px; border-radius:8px; margin-bottom:8px;">
          <div style="font-weight:600; font-size:14px; color:#14532d;">{o['title']}</div>
          <div style="font-size:12px; color:#64748b; margin-top:4px;">{o['desc']}</div>
        </div>"""
    if not opps:
        opps_html = "<div style='font-size:13px; color:#94a3b8; padding:8px 0;'>특별한 기회 신호 없음</div>"

    # AI 액션 플랜
    actions_html = ""
    for i, a in enumerate(ai.get("actions", [])[:3]):
        actions_html += f"""
        <li style="background:#eff6ff; padding:14px 16px 14px 50px; border-radius:8px; margin-bottom:8px; position:relative; list-style:none;">
          <span style="position:absolute; left:14px; top:14px; width:26px; height:26px; background:#3b82f6; color:white; border-radius:50%; display:flex; align-items:center; justify-content:center; font-weight:800; font-size:12px;">{i+1}</span>
          <div style="font-weight:600; font-size:14px; color:#1e3a8a;">{a.get('title', '')}</div>
          <div style="font-size:12px; color:#64748b; margin-top:4px;">{a.get('desc', '')}</div>
        </li>"""
    if not actions_html:
        actions_html = "<div style='font-size:13px; color:#94a3b8; padding:8px 0;'>제안된 액션 없음</div>"

    return f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8">
<style>
@import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard/dist/web/static/pretendard.css');
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Pretendard',-apple-system,sans-serif; background:#f1f5f9; color:#0f172a; line-height:1.6; }}
.container {{ max-width:720px; margin:0 auto; padding:20px; }}
.card {{ background:white; border-radius:16px; box-shadow:0 1px 3px rgba(0,0,0,0.06); overflow:hidden; margin-bottom:16px; }}
</style></head><body>
<div class="container">

<!-- Header -->
<div class="card" style="background:linear-gradient(135deg,#0a0f1c 0%,#1e293b 60%,#1e3a5f 100%); color:white; position:relative; overflow:hidden;">
  <div style="position:absolute; top:-150px; right:-100px; width:400px; height:400px; background:radial-gradient(circle,rgba(59,130,246,0.18) 0%,transparent 70%); pointer-events:none;"></div>
  <div style="position:relative; padding:36px 32px 28px 32px;">
    <div style="display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:28px;">
      <div>
        <div style="color:#60a5fa; font-size:10px; font-weight:800; letter-spacing:2.5px; margin-bottom:10px;">DAILY EXECUTIVE BRIEFING</div>
        <h1 style="font-size:38px; font-weight:900; letter-spacing:-1.5px; line-height:0.95; margin-bottom:8px;">{prefix_en}</h1>
        <div style="color:#03c75a; font-size:18px; font-weight:700; letter-spacing:-0.3px;">{prefix} · {report_type} 네이버 리포트</div>
      </div>
      <div style="background:rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.12); padding:8px 16px; border-radius:20px; font-size:12px; font-weight:600; color:#cbd5e1; white-space:nowrap;">📅 {date_label}</div>
    </div>
    <div style="background:rgba(255,255,255,0.04); border-left:3px solid #3b82f6; padding:20px 22px; border-radius:8px;">
      <div style="font-size:10px; color:#93c5fd; font-weight:700; letter-spacing:1.5px; margin-bottom:10px;">💬 {oneliner_label}</div>
      <div style="font-size:16px; font-weight:500; line-height:1.6;">{ai.get("one_liner", "")}</div>
    </div>
  </div>
</div>

<!-- Key Metrics -->
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b; margin-bottom:16px;">⚡ KEY METRICS</div>
  <div style="display:grid; grid-template-columns:repeat(2,1fr); gap:12px;">
    <div style="background:linear-gradient(135deg,#f0f9ff 0%,#e0f2fe 100%); border:1px solid #bfdbfe; padding:18px; border-radius:12px;">
      <div style="display:flex; align-items:center; gap:6px; margin-bottom:8px;"><span style="font-size:18px;">💰</span><span style="font-size:11px; font-weight:600; color:#1e40af;">매출</span></div>
      <div style="font-size:24px; font-weight:800;">{format_curr(stats['total_revenue'])}</div>
      <div style="margin-top:6px;">{get_delta_chip(stats['total_revenue'], prev_stats['total_revenue'])}</div>
    </div>
    <div style="background:linear-gradient(135deg,#f0fdf4 0%,#dcfce7 100%); border:1px solid #86efac; padding:18px; border-radius:12px;">
      <div style="display:flex; align-items:center; gap:6px; margin-bottom:8px;"><span style="font-size:18px;">🎯</span><span style="font-size:11px; font-weight:600; color:#15803d;">ROAS</span></div>
      <div style="font-size:24px; font-weight:800;">{roas:.0f}%</div>
      <div style="margin-top:6px;">{get_delta_chip(roas, prev_roas, is_percent=True)}</div>
    </div>
    <div style="background:linear-gradient(135deg,#ecfeff 0%,#cffafe 100%); border:1px solid #67e8f9; padding:18px; border-radius:12px;">
      <div style="display:flex; align-items:center; gap:6px; margin-bottom:8px;"><span style="font-size:18px;">💵</span><span style="font-size:11px; font-weight:600; color:#0e7490;">실매출</span></div>
      <div style="font-size:24px; font-weight:800;">{format_curr(net_rev)}</div>
      <div style="margin-top:6px;">{get_delta_chip(net_rev, prev_net)}</div>
      <div style="font-size:10px; color:#64748b; margin-top:4px;">매출 - 환불</div>
    </div>
    <div style="background:linear-gradient(135deg,#fef2f2 0%,#fee2e2 100%); border:1px solid #fca5a5; padding:18px; border-radius:12px;">
      <div style="display:flex; align-items:center; gap:6px; margin-bottom:8px;"><span style="font-size:18px;">↩️</span><span style="font-size:11px; font-weight:600; color:#991b1b;">환불</span></div>
      <div style="font-size:24px; font-weight:800;">{format_curr(stats.get('total_returns', 0))}</div>
      <div style="margin-top:6px;">{get_delta_chip(stats.get('total_returns', 0), prev_stats.get('total_returns', 0), inverse=True)}</div>
      <div style="font-size:10px; color:#64748b; margin-top:4px;">취소 + 반품</div>
    </div>
  </div>
  {cost_kpi_html}
  <div style="background:#f8fafc; padding:10px 12px; border-radius:8px; margin-top:8px; display:flex; justify-content:space-between; align-items:center;">
    <span style="font-size:11px; color:#64748b;">📦 판매 <strong style="color:#0f172a;">{int(stats['total_purchases'] or 0)}건 · 광고비 {format_curr(stats['total_cost'])}</strong></span>
  </div>
</div>

<!-- Bestsellers -->
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">🏆 BESTSELLERS</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:16px;">베스트셀러 TOP 8</h2>
  <div style="display:flex; flex-direction:column; gap:8px;">{bestseller_html}</div>
</div>

<!-- Categories -->
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">📦 CATEGORY MIX</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:16px;">카테고리별 매출 비중</h2>
  {cat_html}
</div>

<!-- Lines (NEW) -->
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">🏷️ PRODUCT LINES</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:16px;">시리즈별 매출 비중</h2>
  {line_html}
</div>

<!-- Hourly (일간만) / Weekly / Monthly 차별 섹션 -->
{hourly_section_html}
{weekly_sections_html}
{monthly_sections_html}

<!-- Channels -->
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">📊 CHANNEL PERFORMANCE</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:16px;">채널별 성과</h2>
  <table style="width:100%; font-size:13px; border-collapse:collapse;">
    <thead><tr style="border-bottom:2px solid #e2e8f0; color:#64748b; font-size:11px; font-weight:600;">
      <th style="text-align:left; padding:8px 6px;">채널</th><th style="text-align:right; padding:8px 6px;">클릭</th>
      <th style="text-align:right; padding:8px 6px;">CVR</th><th style="text-align:right; padding:8px 6px;">매출</th>
    </tr></thead><tbody>{channel_html}</tbody>
  </table>
</div>

<!-- Keywords -->
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">🔍 SEARCH KEYWORDS</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:16px;">검색 키워드 TOP 10</h2>
  <div style="display:flex; flex-wrap:wrap; gap:8px;">{kw_html}</div>
</div>

<!-- Product × Keyword (NEW) -->
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">🎯 PRODUCT × KEYWORD</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:8px;">상품별 유입 키워드</h2>
  <div style="font-size:12px; color:#64748b; margin-bottom:16px;">베스트셀러 5개의 주요 유입 키워드 (TOP 3)</div>
  <div>{pk_html}</div>
</div>

<!-- Alerts & Actions -->
<div class="card" style="padding:24px;">
  <div style="font-size:11px; font-weight:700; letter-spacing:1.5px; color:#64748b;">🚨 ALERTS & ACTIONS</div>
  <h2 style="font-size:18px; font-weight:700; margin-top:4px; margin-bottom:20px;">위험·기회 신호 + AI 액션 플랜</h2>
  <div style="margin-bottom:20px;">
    <div style="font-size:12px; font-weight:700; color:#ef4444; margin-bottom:10px;">⚠️ 위험 신호</div>
    {risks_html}
  </div>
  <div style="margin-bottom:20px;">
    <div style="font-size:12px; font-weight:700; color:#10b981; margin-bottom:10px;">✨ 기회 신호</div>
    {opps_html}
  </div>
  <div>
    <div style="font-size:12px; font-weight:700; color:#3b82f6; margin-bottom:10px;">🎯 AI 액션 플랜 ({action_period})</div>
    <ol style="list-style:none; padding:0;">{actions_html}</ol>
  </div>
</div>

<!-- Footer -->
<div style="text-align:center; padding:24px 0; color:#94a3b8; font-size:11px;">
  <div style="font-weight:600; margin-bottom:4px;">TRAVENCE BIZRAW</div>
  <div>{datetime.now(KST).strftime("%Y-%m-%d %H:%M")} KST 자동 생성</div>
</div>

</div></body></html>"""


# =========================================================
# 6. 메일 발송
# =========================================================
def send_email(subject, html, recipients, sender, gmail_pw):
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = sender
        msg['To'] = ", ".join(recipients)
        msg.attach(MIMEText(html, 'html'))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender, gmail_pw)
            server.sendmail(sender, recipients, msg.as_string())
        return True
    except Exception as e:
        logger.error(f"발송 실패: {e}")
        return False


# =========================================================
# 7. 메인 실행 (브랜드별 진입점에서 호출)
# =========================================================
def _calc_date_range(today, mode):
    """
    모드별로 (현재 기간 시작, 끝, 비교 기간 시작, 끝, 라벨) 반환
    """
    if mode == "weekly":
        # 지난주 (월~일) vs 그 전주
        # today.weekday(): 월=0, 일=6
        # 가장 최근 일요일을 찾음
        days_to_last_sunday = (today.weekday() + 1) % 7
        if days_to_last_sunday == 0:
            days_to_last_sunday = 7  # 오늘이 일요일이면 어제 일요일이 아닌 7일전 일요일
        last_sunday = today - timedelta(days=days_to_last_sunday)
        last_monday = last_sunday - timedelta(days=6)
        # 비교: 그 전주 월~일
        prev_sunday = last_monday - timedelta(days=1)
        prev_monday = prev_sunday - timedelta(days=6)
        return (last_monday, last_sunday, prev_monday, prev_sunday, "전주")

    elif mode == "monthly":
        # 지난달 1일~말일 vs 그 전달
        first_of_this_month = today.replace(day=1)
        last_of_prev_month = first_of_this_month - timedelta(days=1)
        first_of_prev_month = last_of_prev_month.replace(day=1)
        # 비교: 그 전달
        last_of_pprev_month = first_of_prev_month - timedelta(days=1)
        first_of_pprev_month = last_of_pprev_month.replace(day=1)
        return (first_of_prev_month, last_of_prev_month,
                first_of_pprev_month, last_of_pprev_month, "전월")

    else:  # daily
        yesterday = today - timedelta(days=1)
        day_before = today - timedelta(days=2)
        return (yesterday, yesterday, day_before, day_before, "전일")


def run_report(prefix, prefix_en, naver_id, naver_secret, anthropic_key,
               gmail_pw, sender, recipients, save_html_path=None,
               days_offset=0, mode="daily"):
    """
    브랜드별 진입점에서 호출.
    save_html_path가 주어지면 메일 발송 대신 파일로 저장 (로컬 테스트용).
    days_offset: 기준일을 N일 과거로 (자정 직후 어제 데이터 미집계 시 유용)
    mode: "daily" | "weekly" | "monthly"
    """
    mode_label = {"daily": "일간", "weekly": "주간", "monthly": "월간"}[mode]
    logger.info(f"🚀 {prefix} {mode_label} 리포트 시작")

    # === 브랜드별 설정 ===
    brand_config = {
        "팩세이프":   {"show_margin": True,  "catalog": "네이버_팩세이프_마스터_상품명.xlsx"},
        "프레지던트": {"show_margin": False, "catalog": "네이버_프레지던트_마스터_상품명.xlsx"},
    }
    cfg = brand_config.get(prefix, {"show_margin": True, "catalog": ""})
    show_margin = cfg["show_margin"]

    # === 마스터 카탈로그 로드 ===
    script_dir = os.path.dirname(os.path.abspath(__file__))
    catalog_path = os.path.join(script_dir, cfg["catalog"])
    catalog = load_product_catalog(catalog_path)

    token, cno = get_token_and_channel(naver_id, naver_secret)
    if not token or not cno:
        logger.error("토큰/채널 조회 실패. 종료")
        return False
    logger.info(f"  채널: {cno}")

    # === 날짜 범위 계산 ===
    today = datetime.now(KST).date() - timedelta(days=days_offset)
    curr_start, curr_end, prev_start, prev_end, prev_label = _calc_date_range(today, mode)

    curr_start_s = curr_start.strftime("%Y-%m-%d")
    curr_end_s = curr_end.strftime("%Y-%m-%d")
    prev_start_s = prev_start.strftime("%Y-%m-%d")
    prev_end_s = prev_end.strftime("%Y-%m-%d")

    if mode == "daily":
        logger.info(f"📥 데이터 수집: {curr_start_s}")
    else:
        logger.info(f"📥 데이터 수집: {curr_start_s} ~ {curr_end_s} ({(curr_end - curr_start).days + 1}일치)")
    curr = analyze_data(collect_data(token, cno, curr_start_s, curr_end_s), catalog)
    curr["show_margin"] = show_margin

    if mode == "daily":
        logger.info(f"📥 비교 데이터: {prev_start_s}")
    else:
        logger.info(f"📥 비교 데이터: {prev_start_s} ~ {prev_end_s}")
    prev = analyze_data(collect_data(token, cno, prev_start_s, prev_end_s), catalog)
    prev["show_margin"] = show_margin

    # === 주간/월간: 일별 분할 수집 (트렌드 분석용) ===
    daily_breakdown = []
    if mode in ("weekly", "monthly"):
        logger.info(f"📊 일별 분할 수집 시작 ({(curr_end - curr_start).days + 1}일)...")
        d = curr_start
        while d <= curr_end:
            ds = d.strftime("%Y-%m-%d")
            day_data = analyze_data(collect_data(token, cno, ds, ds), catalog)
            net_rev = day_data['total_revenue'] - day_data.get('total_cancel', 0) - day_data['total_refund']
            roas = (net_rev / day_data['total_cost'] * 100) if day_data['total_cost'] > 0 else 0
            daily_breakdown.append({
                'date': d,
                'revenue': day_data['total_revenue'],
                'net_revenue': net_rev,
                'cost': day_data['total_cost'],
                'roas': roas,
                'purchases': day_data['total_purchases'],
                'gp': day_data.get('total_gp', 0),
                'top_product': day_data['products'][0].get('_master_name', '?')[:30] if day_data['products'] else '?',
                'top_keyword': day_data['keywords'][0].get('refKeyword', '?') if day_data['keywords'] else '?',
            })
            d += timedelta(days=1)
        curr['daily_breakdown'] = daily_breakdown
        logger.info(f"  ✓ 일별 데이터 {len(daily_breakdown)}일치 수집 완료")

        # 베스트셀러 순위 변동 (현재 vs 이전 기간)
        prev_ranks = {p.get('_master_name', p.get('productName', '')): i for i, p in enumerate(prev['products'][:30])}
        for i, p in enumerate(curr['products'][:10]):
            name = p.get('_master_name') or p.get('productName', '')
            prev_rank = prev_ranks.get(name)
            if prev_rank is not None:
                p['_rank_change'] = prev_rank - i  # 양수=상승, 음수=하락
                p['_prev_rank'] = prev_rank + 1
            else:
                p['_rank_change'] = None  # 신규 진입
                p['_prev_rank'] = None

        # 카테고리/시리즈 순위 변동
        prev_cat_ranks = {c['name']: i for i, c in enumerate(prev.get('categories', []))}
        for i, c in enumerate(curr.get('categories', [])):
            prev_rank = prev_cat_ranks.get(c['name'])
            c['rank_change'] = (prev_rank - i) if prev_rank is not None else None

        prev_line_ranks = {l['name']: i for i, l in enumerate(prev.get('lines', []))}
        for i, l in enumerate(curr.get('lines', [])):
            prev_rank = prev_line_ranks.get(l['name'])
            l['rank_change'] = (prev_rank - i) if prev_rank is not None else None

        # 키워드 변화 (현재 vs 이전)
        prev_kw_ranks = {k.get('refKeyword'): i for i, k in enumerate(prev.get('keywords', [])[:30])}
        for i, k in enumerate(curr.get('keywords', [])[:10]):
            kw_text = k.get('refKeyword')
            prev_rank = prev_kw_ranks.get(kw_text)
            k['rank_change'] = (prev_rank - i) if prev_rank is not None else None

    logger.info(f"🔍 위험/기회 신호 탐지 중...")
    risks, opps = detect_alerts(curr, prev, prev_label=prev_label)

    logger.info(f"🤖 AI 분석 중 (Opus 4.7)...")
    ai = get_ai_analysis(curr, prev, prefix, mode_label, anthropic_key)

    logger.info(f"🎨 HTML 생성 중...")
    html = render_html(prefix, prefix_en, curr, prev, ai, risks, opps,
                       curr_start_s, curr_end_s, mode_label)

    if save_html_path:
        with open(save_html_path, "w", encoding="utf-8") as f:
            f.write(html)
        logger.info(f"💾 저장: {save_html_path}")
        return True

    # 메일 제목 (모드별)
    if mode == "daily":
        subject = f"📊 [{prefix}] 일간 네이버 리포트 ({curr_start_s})"
    elif mode == "weekly":
        subject = f"📊 [{prefix}] 주간 네이버 리포트 ({curr_start_s} ~ {curr_end_s})"
    else:  # monthly
        subject = f"📊 [{prefix}] 월간 네이버 리포트 ({curr_start.strftime('%Y년 %m월')})"

    if send_email(subject, html, recipients, sender, gmail_pw):
        logger.info(f"✅ 발송 성공!")
        return True
    return False
